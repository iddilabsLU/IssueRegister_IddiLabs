"""Export service for Excel operations."""

import shutil
import tempfile
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.database.models import Issue, User, Status, RiskLevel
from src.database.connection import DatabaseConnection


class ExportService:
    """
    Handles data export and import operations.

    Supports:
    - Excel export of issues
    - Excel import (bulk)
    - Database backup/restore
    - Import template generation
    """

    # Column configuration for export
    EXPORT_COLUMNS = [
        ("ID", "id", 8),
        ("Title", "title", 40),
        ("Status", "status", 15),
        ("Summary", "summary_description", 50),
        ("Topic", "topic", 20),
        ("Identified By", "identified_by", 20),
        ("Owner", "owner", 20),
        ("Department", "department", 15),
        ("Description", "description", 60),
        ("Remediation Action", "remediation_action", 60),
        ("Risk Description", "risk_description", 40),
        ("Risk Level", "risk_level", 12),
        ("Identification Date", "identification_date", 15),
        ("Due Date", "due_date", 15),
        ("Follow-up Date", "follow_up_date", 15),
        ("Updates", "updates", 60),
        ("Closing Date", "closing_date", 15),
    ]

    # Header style
    HEADER_FILL = PatternFill(start_color="2D3E50", end_color="2D3E50", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True)
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

    # Cell border
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    def export_issues_to_excel(
        self,
        issues: list[Issue],
        file_path: str
    ) -> tuple[bool, str]:
        """
        Export issues to an Excel file.

        Args:
            issues: List of issues to export
            file_path: Path to save the Excel file

        Returns:
            Tuple of (success, error_message)
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Issues"

            # Write headers
            for col_idx, (header, _, width) in enumerate(self.EXPORT_COLUMNS, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = self.HEADER_FILL
                cell.font = self.HEADER_FONT
                cell.alignment = self.HEADER_ALIGNMENT
                cell.border = self.THIN_BORDER
                ws.column_dimensions[get_column_letter(col_idx)].width = width

            # Write data
            for row_idx, issue in enumerate(issues, 2):
                for col_idx, (_, field, _) in enumerate(self.EXPORT_COLUMNS, 1):
                    value = getattr(issue, field, None)

                    # Format dates
                    if value is not None and field.endswith("_date"):
                        value = value.isoformat() if hasattr(value, "isoformat") else str(value)

                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = self.THIN_BORDER
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

            # Freeze header row
            ws.freeze_panes = "A2"

            wb.save(file_path)
            return True, ""

        except Exception as e:
            return False, f"Export failed: {str(e)}"

    def create_import_template(self, file_path: str) -> tuple[bool, str]:
        """
        Create an Excel template for bulk import.

        Args:
            file_path: Path to save the template

        Returns:
            Tuple of (success, error_message)
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Import Template"

            # Template columns (exclude ID - auto-generated)
            template_columns = [col for col in self.EXPORT_COLUMNS if col[1] != "id"]

            # Write headers
            for col_idx, (header, _, width) in enumerate(template_columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = self.HEADER_FILL
                cell.font = self.HEADER_FONT
                cell.alignment = self.HEADER_ALIGNMENT
                cell.border = self.THIN_BORDER
                ws.column_dimensions[get_column_letter(col_idx)].width = width

            # Add example row
            example_data = {
                "title": "Example Issue Title",
                "status": "Open",
                "summary_description": "Brief description of the issue",
                "topic": "System Error",
                "identified_by": "John Smith",
                "owner": "Jane Doe",
                "department": "IT",
                "description": "Detailed description of the issue...",
                "remediation_action": "Steps to fix the issue...",
                "risk_description": "Potential impact if not resolved",
                "risk_level": "Medium",
                "identification_date": datetime.now().strftime("%Y-%m-%d"),
                "due_date": "",
                "follow_up_date": "",
                "updates": "",
                "closing_date": "",
            }

            for col_idx, (_, field, _) in enumerate(template_columns, 1):
                value = example_data.get(field, "")
                cell = ws.cell(row=2, column=col_idx, value=value)
                cell.border = self.THIN_BORDER
                cell.alignment = Alignment(vertical="top")

            # Add instructions sheet
            ws_help = wb.create_sheet("Instructions")
            instructions = [
                ("Field", "Description", "Valid Values"),
                ("Title", "Brief title for the issue (required)", "Any text"),
                ("Status", "Current status of the issue", "Draft, Open, In Progress, Remediated, Closed"),
                ("Summary", "Short summary description", "Any text"),
                ("Topic", "Category of the issue", "Any text (e.g., System Error, Policy Violation)"),
                ("Identified By", "Person who found the issue", "Any text"),
                ("Owner", "Person responsible for resolution", "Any text"),
                ("Department", "Related department", "Any text"),
                ("Description", "Detailed issue description", "Any text"),
                ("Remediation Action", "Actions to resolve", "Any text"),
                ("Risk Description", "Description of risk/impact", "Any text"),
                ("Risk Level", "Severity of the risk", "None, Low, Medium, High"),
                ("Identification Date", "Date issue was found", "YYYY-MM-DD format"),
                ("Due Date", "Target resolution date", "YYYY-MM-DD format"),
                ("Follow-up Date", "Next review date", "YYYY-MM-DD format"),
                ("Updates", "Progress notes", "Any text"),
                ("Closing Date", "Date issue was closed", "YYYY-MM-DD format"),
            ]

            for row_idx, (col1, col2, col3) in enumerate(instructions, 1):
                ws_help.cell(row=row_idx, column=1, value=col1)
                ws_help.cell(row=row_idx, column=2, value=col2)
                ws_help.cell(row=row_idx, column=3, value=col3)

                if row_idx == 1:
                    for col in range(1, 4):
                        cell = ws_help.cell(row=1, column=col)
                        cell.fill = self.HEADER_FILL
                        cell.font = self.HEADER_FONT

            ws_help.column_dimensions["A"].width = 20
            ws_help.column_dimensions["B"].width = 40
            ws_help.column_dimensions["C"].width = 50

            wb.save(file_path)
            return True, ""

        except Exception as e:
            return False, f"Template creation failed: {str(e)}"

    def import_issues_from_excel(
        self,
        file_path: str
    ) -> tuple[int, list[str]]:
        """
        Import issues from an Excel file.

        Args:
            file_path: Path to the Excel file

        Returns:
            Tuple of (success_count, list_of_errors)
        """
        from src.database import queries

        errors = []
        success_count = 0

        try:
            wb = load_workbook(file_path)
            ws = wb.active

            # Get header row to map columns
            headers = {}
            for col_idx, cell in enumerate(ws[1], 1):
                if cell.value:
                    headers[cell.value.lower().replace(" ", "_").replace("-", "_")] = col_idx

            # Process data rows
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
                try:
                    # Extract values
                    def get_value(field_name):
                        col = headers.get(field_name)
                        if col:
                            return row[col - 1].value
                        return None

                    title = get_value("title")
                    if not title:
                        errors.append(f"Row {row_idx}: Missing title (required)")
                        continue

                    # Parse and validate status
                    status = get_value("status") or "Draft"
                    if status not in Status.values():
                        status = "Draft"

                    # Parse and validate risk level
                    risk_level = get_value("risk_level") or "None"
                    if risk_level not in RiskLevel.values():
                        risk_level = "None"

                    # Parse dates
                    def parse_date(field_name):
                        value = get_value(field_name)
                        if value is None:
                            return None
                        if hasattr(value, "date"):
                            return value.date()
                        if isinstance(value, str):
                            try:
                                from datetime import date
                                return date.fromisoformat(value)
                            except ValueError:
                                return None
                        return None

                    issue = Issue(
                        title=title,
                        status=status,
                        summary_description=get_value("summary") or get_value("summary_description"),
                        topic=get_value("topic"),
                        identified_by=get_value("identified_by"),
                        owner=get_value("owner"),
                        department=get_value("department"),
                        description=get_value("description"),
                        remediation_action=get_value("remediation_action"),
                        risk_description=get_value("risk_description"),
                        risk_level=risk_level,
                        identification_date=parse_date("identification_date"),
                        due_date=parse_date("due_date"),
                        follow_up_date=parse_date("follow_up_date"),
                        updates=get_value("updates"),
                        closing_date=parse_date("closing_date"),
                        supporting_docs=[],
                    )

                    queries.create_issue(issue)
                    success_count += 1

                except Exception as e:
                    errors.append(f"Row {row_idx}: {str(e)}")

        except Exception as e:
            errors.append(f"Failed to read file: {str(e)}")

        return success_count, errors

    def backup_database(self, backup_path: str) -> tuple[bool, str]:
        """
        Create a backup of the current database and attachments.

        Creates a ZIP file containing:
        - Database file (.db, -wal, -shm)
        - Attachments folder (excluding _staging, including _deleted)

        Args:
            backup_path: Path to save the backup ZIP file

        Returns:
            Tuple of (success, error_message)
        """
        try:
            db = DatabaseConnection.get_instance()
            source_db_path = db.db_path

            # Ensure we flush any pending changes
            db.commit()

            # Get attachments folder path
            attachments_root = source_db_path.parent / "attachments"

            # Create ZIP file
            with zipfile.ZipFile(backup_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                # Add database file
                zipf.write(str(source_db_path), "database.db")

                # Add WAL and SHM files if they exist
                for suffix in ["-wal", "-shm"]:
                    src = str(source_db_path) + suffix
                    if Path(src).exists():
                        zipf.write(src, f"database.db{suffix}")

                # Add attachments folder (if it exists)
                if attachments_root.exists():
                    for item in attachments_root.rglob("*"):
                        # Skip _staging folder entirely
                        if "_staging" in item.parts:
                            continue

                        # Add files to ZIP maintaining folder structure
                        if item.is_file():
                            # Get path relative to attachments root
                            rel_path = item.relative_to(attachments_root)
                            zipf.write(str(item), f"attachments/{rel_path}")

            return True, ""

        except Exception as e:
            return False, f"Backup failed: {str(e)}"

    def restore_database(self, backup_path: str) -> tuple[bool, str]:
        """
        Restore database and attachments from a backup ZIP file.

        WARNING: This overwrites the current database and attachments!

        Args:
            backup_path: Path to the backup ZIP file

        Returns:
            Tuple of (success, error_message)
        """
        temp_dir = None
        try:
            if not Path(backup_path).exists():
                return False, "Backup file not found."

            # Check if it's a ZIP file
            if not zipfile.is_zipfile(backup_path):
                return False, "Backup file is not a valid ZIP file."

            db = DatabaseConnection.get_instance()
            target_db_path = db.db_path
            target_db_dir = target_db_path.parent

            # Close current connection
            db.close()

            # Create temporary directory for extraction
            temp_dir = tempfile.mkdtemp(prefix="issue_register_restore_")
            temp_path = Path(temp_dir)

            # Extract ZIP to temporary location
            with zipfile.ZipFile(backup_path, 'r') as zipf:
                zipf.extractall(temp_path)

            # Verify extracted database exists
            extracted_db = temp_path / "database.db"
            if not extracted_db.exists():
                return False, "Backup ZIP does not contain database.db file."

            # Remove existing database files
            for suffix in ["", "-wal", "-shm"]:
                path = str(target_db_path) + suffix
                if Path(path).exists():
                    Path(path).unlink()

            # Copy database files to target location
            shutil.copy2(str(extracted_db), str(target_db_path))

            # Copy WAL and SHM if they exist in backup
            for suffix in ["-wal", "-shm"]:
                src = temp_path / f"database.db{suffix}"
                if src.exists():
                    shutil.copy2(str(src), str(target_db_path) + suffix)

            # Restore attachments folder if it exists in backup
            extracted_attachments = temp_path / "attachments"
            if extracted_attachments.exists():
                target_attachments = target_db_dir / "attachments"

                # Remove existing attachments folder (except _staging to preserve unsaved work)
                if target_attachments.exists():
                    # Preserve staging folder if it exists
                    staging_backup = None
                    staging_folder = target_attachments / "_staging"
                    if staging_folder.exists():
                        staging_backup = tempfile.mkdtemp(prefix="staging_backup_")
                        shutil.copytree(str(staging_folder), str(Path(staging_backup) / "_staging"))

                    # Remove attachments folder
                    shutil.rmtree(str(target_attachments), ignore_errors=True)

                    # Restore staging if it was backed up
                    if staging_backup:
                        target_attachments.mkdir(parents=True, exist_ok=True)
                        shutil.copytree(
                            str(Path(staging_backup) / "_staging"),
                            str(target_attachments / "_staging")
                        )
                        shutil.rmtree(staging_backup, ignore_errors=True)

                # Copy attachments from backup
                shutil.copytree(str(extracted_attachments), str(target_attachments), dirs_exist_ok=True)

            # Clean up temporary directory
            shutil.rmtree(temp_dir, ignore_errors=True)
            temp_dir = None

            # Reconnect to database
            DatabaseConnection.reset_instance()
            DatabaseConnection.get_instance(str(target_db_path))

            return True, ""

        except Exception as e:
            # Clean up temporary directory on error
            if temp_dir and Path(temp_dir).exists():
                shutil.rmtree(temp_dir, ignore_errors=True)
            return False, f"Restore failed: {str(e)}"

    def export_users_to_excel(self, file_path: str) -> tuple[bool, str]:
        """
        Export users list to an Excel file.

        Args:
            file_path: Path to save the Excel file

        Returns:
            Tuple of (success, error_message)
        """
        from src.database import queries

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Users"

            # Define columns
            columns = [
                ("ID", 8),
                ("Username", 20),
                ("Role", 15),
                ("Departments", 40),
                ("View Departments", 40),
                ("Edit Departments", 40),
            ]

            # Write headers
            for col_idx, (header, width) in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = self.HEADER_FILL
                cell.font = self.HEADER_FONT
                cell.alignment = self.HEADER_ALIGNMENT
                cell.border = self.THIN_BORDER
                ws.column_dimensions[get_column_letter(col_idx)].width = width

            # Write data
            users = queries.list_users()
            for row_idx, user in enumerate(users, 2):
                ws.cell(row=row_idx, column=1, value=user.id).border = self.THIN_BORDER
                ws.cell(row=row_idx, column=2, value=user.username).border = self.THIN_BORDER
                ws.cell(row=row_idx, column=3, value=user.role).border = self.THIN_BORDER
                ws.cell(row=row_idx, column=4, value=", ".join(user.departments) if user.departments else "All").border = self.THIN_BORDER
                ws.cell(row=row_idx, column=5, value=", ".join(user.view_departments) if user.view_departments else "All").border = self.THIN_BORDER
                ws.cell(row=row_idx, column=6, value=", ".join(user.edit_departments) if user.edit_departments else "All").border = self.THIN_BORDER

            # Freeze header row
            ws.freeze_panes = "A2"

            wb.save(file_path)
            return True, ""

        except Exception as e:
            return False, f"Export failed: {str(e)}"

    def export_audit_log(self, file_path: str) -> tuple[bool, str]:
        """
        Export audit log to an Excel file.

        Args:
            file_path: Path to save the Excel file

        Returns:
            Tuple of (success, error_message)
        """
        from src.database import queries
        import json

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Audit Log"

            # Define columns
            columns = [
                ("ID", 8),
                ("Timestamp", 20),
                ("Username", 15),
                ("Action", 15),
                ("Entity Type", 12),
                ("Entity ID", 10),
                ("Details", 80),
            ]

            # Write headers
            for col_idx, (header, width) in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = self.HEADER_FILL
                cell.font = self.HEADER_FONT
                cell.alignment = self.HEADER_ALIGNMENT
                cell.border = self.THIN_BORDER
                ws.column_dimensions[get_column_letter(col_idx)].width = width

            # Write data
            logs = queries.list_audit_logs()
            for row_idx, log in enumerate(logs, 2):
                ws.cell(row=row_idx, column=1, value=log.id).border = self.THIN_BORDER
                ws.cell(row=row_idx, column=2, value=log.timestamp.strftime("%Y-%m-%d %H:%M:%S") if log.timestamp else "").border = self.THIN_BORDER
                ws.cell(row=row_idx, column=3, value=log.username).border = self.THIN_BORDER
                ws.cell(row=row_idx, column=4, value=log.action).border = self.THIN_BORDER
                ws.cell(row=row_idx, column=5, value=log.entity_type).border = self.THIN_BORDER
                ws.cell(row=row_idx, column=6, value=log.entity_id).border = self.THIN_BORDER

                # Format details as readable text
                details_text = ""
                if log.details:
                    if isinstance(log.details, dict):
                        details_text = json.dumps(log.details, indent=2)
                    else:
                        details_text = str(log.details)

                cell = ws.cell(row=row_idx, column=7, value=details_text)
                cell.border = self.THIN_BORDER
                cell.alignment = Alignment(vertical="top", wrap_text=True)

            # Freeze header row
            ws.freeze_panes = "A2"

            wb.save(file_path)
            return True, ""

        except Exception as e:
            return False, f"Export failed: {str(e)}"


# Singleton instance
_export_service: Optional[ExportService] = None


def get_export_service() -> ExportService:
    """Get the singleton ExportService instance."""
    global _export_service
    if _export_service is None:
        _export_service = ExportService()
    return _export_service
